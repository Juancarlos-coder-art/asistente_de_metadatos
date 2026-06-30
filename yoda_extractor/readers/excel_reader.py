from typing import Generator
from .base import BaseReader


class ExcelReader(BaseReader):
    def stream_records(self) -> Generator[dict, None, None]:
        from openpyxl import load_workbook

        wb = load_workbook(self.path, read_only=True, data_only=True)
        for sheet in wb.worksheets:
            # Read first 10 rows to detect the correct header row
            rows_preview = []
            for r in sheet.iter_rows(values_only=True):
                rows_preview.append(r)
                if len(rows_preview) >= 10:
                    break

            if not rows_preview:
                continue

            # Identify the row with the highest number of non-empty cells in the preview
            header_idx = 0
            max_non_empty = -1
            for idx, r in enumerate(rows_preview):
                non_empty = sum(1 for c in r if c is not None and str(c).strip() != "")
                if non_empty > max_non_empty:
                    max_non_empty = non_empty
                    header_idx = idx

            # Construct headers from the selected row
            raw_header = rows_preview[header_idx]
            headers = [
                str(c).strip() if c is not None and str(c).strip() != "" else f"col_{i}"
                for i, c in enumerate(raw_header)
            ]

            # Stream rows starting after the selected header row
            current_row_idx = 0
            for row in sheet.iter_rows(values_only=True):
                if current_row_idx <= header_idx:
                    current_row_idx += 1
                    continue
                record = {}
                for header, cell in zip(headers, row):
                    if cell is None:
                        continue
                    val = str(cell).strip()
                    if val and val not in ("nan", "NaN", "None"):
                        record[header] = val
                if record:
                    yield record
        wb.close()

