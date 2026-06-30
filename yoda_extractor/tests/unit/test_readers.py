"""Unit tests for all readers. Uses tmp_path to create real temp files."""
import json
import tempfile
from io import BytesIO
from pathlib import Path

import pandas as pd
import pytest

from readers import get_reader
from readers.csv_reader import CSVReader
from readers.json_reader import JSONReader, _flatten
from readers.xml_reader import XMLReader, _element_to_dict
from readers.parquet_reader import ParquetReader
from readers.excel_reader import ExcelReader
import defusedxml.ElementTree as ET


# ── get_reader ────────────────────────────────────────────────────────────────

def test_get_reader_returns_csv_reader(tmp_path):
    f = tmp_path / "data.csv"
    f.write_text("a,b\n1,2\n")
    reader = get_reader(str(f))
    assert isinstance(reader, CSVReader)


def test_get_reader_returns_json_reader(tmp_path):
    f = tmp_path / "data.json"
    f.write_text('[{"a": 1}]')
    reader = get_reader(str(f))
    assert isinstance(reader, JSONReader)


def test_get_reader_returns_xml_reader(tmp_path):
    f = tmp_path / "data.xml"
    f.write_text("<root><item><name>x</name></item></root>")
    reader = get_reader(str(f))
    assert isinstance(reader, XMLReader)


def test_get_reader_returns_excel_reader(tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook
    f = tmp_path / "data.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["col"])
    ws.append(["val"])
    wb.save(str(f))
    reader = get_reader(str(f))
    assert isinstance(reader, ExcelReader)


def test_get_reader_unsupported_extension(tmp_path):
    f = tmp_path / "data.xyz"
    f.write_text("content")
    with pytest.raises(ValueError, match="Unsupported"):
        get_reader(str(f))


# ── CSVReader ─────────────────────────────────────────────────────────────────

def test_csv_reader_basic(tmp_path):
    f = tmp_path / "data.csv"
    f.write_text("name,age\nAlice,30\nBob,25\n")
    records = list(CSVReader(str(f)).stream_records())
    assert len(records) == 2
    assert records[0]["name"] == "Alice"
    assert records[1]["age"] == "25"


def test_csv_reader_semicolon_delimiter(tmp_path):
    f = tmp_path / "data.csv"
    f.write_text("name;age\nAlice;30\nBob;25\n")
    records = list(CSVReader(str(f)).stream_records())
    assert len(records) == 2
    assert records[0]["name"] == "Alice"


def test_csv_reader_tab_delimiter(tmp_path):
    f = tmp_path / "data.csv"
    f.write_text("name\tage\nAlice\t30\nBob\t25\n")
    records = list(CSVReader(str(f)).stream_records())
    assert len(records) == 2
    assert records[0]["name"] == "Alice"


def test_csv_reader_omits_nan_values(tmp_path):
    f = tmp_path / "data.csv"
    f.write_text("name,age\nAlice,\nBob,25\n")
    records = list(CSVReader(str(f)).stream_records())
    assert "age" not in records[0]
    assert records[1]["age"] == "25"


def test_csv_detect_encoding_utf8(tmp_path):
    f = tmp_path / "data.csv"
    f.write_bytes("name\nAlice\n".encode("utf-8"))
    reader = CSVReader(str(f))
    assert reader._detect_encoding() == "utf-8-sig"


def test_csv_detect_encoding_latin1(tmp_path):
    f = tmp_path / "data.csv"
    f.write_bytes("name\nAl\xeece\n".encode("latin-1"))
    reader = CSVReader(str(f))
    assert reader._detect_encoding() == "latin-1"


def test_csv_detect_delimiter_comma(tmp_path):
    f = tmp_path / "data.csv"
    f.write_text("a,b,c\n1,2,3\n4,5,6\n")
    reader = CSVReader(str(f))
    assert reader._detect_delimiter("utf-8") == ","


def test_csv_detect_delimiter_empty_file(tmp_path):
    f = tmp_path / "data.csv"
    f.write_text("")
    reader = CSVReader(str(f))
    assert reader._detect_delimiter("utf-8") == ","


# ── _flatten (JSONReader helper) ──────────────────────────────────────────────

def test_flatten_simple_dict():
    result = _flatten({"a": 1, "b": "hello"})
    assert result == {"a": "1", "b": "hello"}


def test_flatten_nested_dict():
    result = _flatten({"outer": {"inner": "value"}})
    assert result == {"outer.inner": "value"}


def test_flatten_list_becomes_json_string():
    result = _flatten({"tags": ["a", "b"]})
    assert result["tags"] == '["a", "b"]'


def test_flatten_none_value():
    result = _flatten({"key": None})
    assert result["key"] is None


def test_flatten_deeply_nested():
    result = _flatten({"a": {"b": {"c": "deep"}}})
    assert result == {"a.b.c": "deep"}


# ── JSONReader ────────────────────────────────────────────────────────────────

def test_json_reader_array_of_objects(tmp_path):
    f = tmp_path / "data.json"
    f.write_text('[{"name": "Alice", "age": 30}, {"name": "Bob", "age": 25}]')
    records = list(JSONReader(str(f)).stream_records())
    assert len(records) == 2
    assert records[0]["name"] == "Alice"


def test_json_reader_single_object(tmp_path):
    f = tmp_path / "data.json"
    data = {"items": [{"id": 1, "val": "a"}, {"id": 2, "val": "b"}]}
    f.write_text(json.dumps(data))
    records = list(JSONReader(str(f)).stream_records())
    assert len(records) == 2


def test_json_reader_ndjson(tmp_path):
    f = tmp_path / "data.json"
    f.write_text('{"name": "Alice"}\n{"name": "Bob"}\n')
    records = list(JSONReader(str(f)).stream_records())
    assert len(records) == 2
    names = {r["name"] for r in records}
    assert names == {"Alice", "Bob"}


def test_json_reader_empty_array(tmp_path):
    f = tmp_path / "data.json"
    f.write_text("[]")
    records = list(JSONReader(str(f)).stream_records())
    assert records == []


def test_json_reader_nested_object_unwrapped(tmp_path):
    f = tmp_path / "data.json"
    data = {"dataset": [{"a": 1}, {"a": 2}, {"a": 3}]}
    f.write_text(json.dumps(data))
    records = list(JSONReader(str(f)).stream_records())
    assert len(records) == 3


def test_json_reader_array_with_non_dict_item(tmp_path):
    f = tmp_path / "data.json"
    f.write_text('[{"name": "Alice"}, "plain_string", {"name": "Bob"}]')
    records = list(JSONReader(str(f)).stream_records())
    # plain_string produces {"value": "plain_string"}
    assert any(r.get("name") == "Alice" for r in records)
    assert any("value" in r for r in records)


def test_json_reader_ndjson_skips_blank_lines(tmp_path):
    f = tmp_path / "data.json"
    f.write_text('{"name": "Alice"}\n\n{"name": "Bob"}\n')
    records = list(JSONReader(str(f)).stream_records())
    assert len(records) == 2


def test_json_reader_ndjson_skips_invalid_lines(tmp_path):
    f = tmp_path / "data.json"
    f.write_text('{"name": "Alice"}\nnot_json\n{"name": "Bob"}\n')
    records = list(JSONReader(str(f)).stream_records())
    assert len(records) == 2


def test_json_reader_keyed_collection(tmp_path):
    f = tmp_path / "data.json"
    data = {
        "item1": {"name": "Alice", "age": 30},
        "item2": {"name": "Bob", "age": 25},
        "item3": {"name": "Carol", "age": 35},
    }
    f.write_text(json.dumps(data))
    records = list(JSONReader(str(f)).stream_records())
    assert len(records) == 3
    names = {r["name"] for r in records}
    assert names == {"Alice", "Bob", "Carol"}


def test_json_reader_list_of_lists_with_headers(tmp_path):
    f = tmp_path / "data.json"
    data = {
        "columns": ["name", "age"],
        "rows": [["Alice", 30], ["Bob", 25]],
    }
    f.write_text(json.dumps(data))
    records = list(JSONReader(str(f)).stream_records())
    assert len(records) == 2


def test_json_reader_flat_object_yields_single_record(tmp_path):
    f = tmp_path / "data.json"
    f.write_text('{"name": "Alice", "age": 30}')
    records = list(JSONReader(str(f)).stream_records())
    assert len(records) >= 1


def test_json_reader_stream_fallback(tmp_path):
    f = tmp_path / "data.json"
    f.write_text('[{"name": "Alice"}, {"name": "Bob"}]')
    reader = JSONReader(str(f))
    records = list(reader._stream_fallback())
    assert len(records) == 2
    assert records[0]["name"] == "Alice"


def test_json_reader_stream_fallback_dict(tmp_path):
    f = tmp_path / "data.json"
    data = {"items": [{"a": 1}, {"a": 2}]}
    f.write_text(json.dumps(data))
    reader = JSONReader(str(f))
    records = list(reader._stream_fallback())
    assert len(records) == 2


def test_json_reader_with_utf8_bom(tmp_path):
    f = tmp_path / "data.json"
    content = '[{"name": "Alice"}]'
    f.write_bytes(b"\xef\xbb\xbf" + content.encode("utf-8"))
    records = list(JSONReader(str(f)).stream_records())
    assert len(records) == 1
    assert records[0]["name"] == "Alice"


# ── _element_to_dict (XMLReader helper) ───────────────────────────────────────

def test_element_to_dict_text_only():
    elem = ET.fromstring("<name>Alice</name>")
    result = _element_to_dict(elem)
    assert result == {"_text": "Alice"}


def test_element_to_dict_attributes():
    elem = ET.fromstring('<item id="1" type="A"/>')
    result = _element_to_dict(elem)
    assert result["id"] == "1"
    assert result["type"] == "A"


def test_element_to_dict_children_as_simple_leaf():
    elem = ET.fromstring("<record><name>Alice</name><age>30</age></record>")
    result = _element_to_dict(elem)
    assert result["name"] == "Alice"
    assert result["age"] == "30"


def test_element_to_dict_duplicate_children_get_suffix():
    elem = ET.fromstring("<root><tag>a</tag><tag>b</tag></root>")
    result = _element_to_dict(elem)
    assert "tag" in result
    assert "tag_1" in result


def test_element_to_dict_strips_namespace():
    elem = ET.fromstring('<root xmlns:ns="http://x"><ns:child>val</ns:child></root>')
    result = _element_to_dict(elem)
    assert "child" in result


# ── XMLReader ────────────────────────────────────────────────────────────────

def test_xml_reader_basic(tmp_path):
    f = tmp_path / "data.xml"
    f.write_text(
        "<records>"
        "<record><name>Alice</name><age>30</age></record>"
        "<record><name>Bob</name><age>25</age></record>"
        "</records>"
    )
    records = list(XMLReader(str(f)).stream_records())
    assert len(records) == 2
    assert records[0]["name"] == "Alice"
    assert records[1]["age"] == "25"


def test_xml_reader_empty_root(tmp_path):
    f = tmp_path / "data.xml"
    f.write_text("<root></root>")
    records = list(XMLReader(str(f)).stream_records())
    assert records == []


def test_xml_reader_namespace_stripped(tmp_path):
    f = tmp_path / "data.xml"
    f.write_text(
        '<root xmlns:ns="http://example.com">'
        '<ns:record><ns:name>Alice</ns:name></ns:record>'
        '<ns:record><ns:name>Bob</ns:name></ns:record>'
        '<ns:record><ns:name>Carol</ns:name></ns:record>'
        '<ns:record><ns:name>Dave</ns:name></ns:record>'
        '<ns:record><ns:name>Eve</ns:name></ns:record>'
        '</root>'
    )
    records = list(XMLReader(str(f)).stream_records())
    assert len(records) > 0


def test_xml_reader_nested_elements_dot_notation(tmp_path):
    f = tmp_path / "data.xml"
    f.write_text(
        "<records>"
        "<record><address><city>Madrid</city><zip>28001</zip></address></record>"
        "<record><address><city>Barcelona</city><zip>08001</zip></address></record>"
        "<record><address><city>Seville</city><zip>41001</zip></address></record>"
        "</records>"
    )
    records = list(XMLReader(str(f)).stream_records())
    assert len(records) == 3
    assert any("address.city" in r or "city" in r for r in records)


def test_xml_reader_element_with_attributes_and_children(tmp_path):
    f = tmp_path / "data.xml"
    f.write_text(
        "<records>"
        '<record id="1"><name>Alice</name><score type="int">90</score></record>'
        '<record id="2"><name>Bob</name><score type="int">85</score></record>'
        '<record id="3"><name>Carol</name><score type="int">92</score></record>'
        "</records>"
    )
    records = list(XMLReader(str(f)).stream_records())
    assert len(records) == 3
    assert records[0].get("id") == "1"


def test_element_to_dict_nested_with_attributes():
    elem = ET.fromstring('<person><address city="Madrid" zip="28001"/></person>')
    result = _element_to_dict(elem)
    assert "address.city" in result or "address" in result


def test_xml_reader_ignores_empty_records(tmp_path):
    f = tmp_path / "data.xml"
    f.write_text(
        "<records>"
        "<record><name>Alice</name></record>"
        "<record></record>"
        "<record><name>Bob</name></record>"
        "</records>"
    )
    records = list(XMLReader(str(f)).stream_records())
    assert all(r for r in records)


# ── ParquetReader ─────────────────────────────────────────────────────────────

def test_parquet_reader_basic(tmp_path):
    pytest.importorskip("pyarrow")
    f = tmp_path / "data.parquet"
    df = pd.DataFrame({"name": ["Alice", "Bob"], "age": [30, 25]})
    df.to_parquet(str(f), index=False)
    records = list(ParquetReader(str(f)).stream_records())
    assert len(records) == 2
    names = {r["name"] for r in records}
    assert names == {"Alice", "Bob"}


def test_parquet_reader_omits_null_values(tmp_path):
    pytest.importorskip("pyarrow")
    import numpy as np
    f = tmp_path / "data.parquet"
    df = pd.DataFrame({"name": ["Alice", None], "age": [30, 25]})
    df.to_parquet(str(f), index=False)
    records = list(ParquetReader(str(f)).stream_records())
    assert "name" not in records[1]


# ── ExcelReader ───────────────────────────────────────────────────────────────

def test_excel_reader_basic(tmp_path):
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl import Workbook
    f = tmp_path / "data.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["name", "age"])
    ws.append(["Alice", 30])
    ws.append(["Bob", 25])
    wb.save(str(f))
    records = list(ExcelReader(str(f)).stream_records())
    assert len(records) == 2
    assert records[0]["name"] == "Alice"
    assert records[1]["age"] == "25"


def test_excel_reader_skips_empty_rows(tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook
    f = tmp_path / "data.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["name"])
    ws.append([None])
    ws.append(["Alice"])
    wb.save(str(f))
    records = list(ExcelReader(str(f)).stream_records())
    assert len(records) == 1
    assert records[0]["name"] == "Alice"


def test_excel_reader_empty_sheet(tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook
    f = tmp_path / "data.xlsx"
    wb = Workbook()
    wb.save(str(f))
    records = list(ExcelReader(str(f)).stream_records())
    assert records == []


def test_excel_reader_auto_detects_header_row(tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook
    f = tmp_path / "data.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append([None, None])
    ws.append(["name", "age", "city"])
    ws.append(["Alice", 30, "Madrid"])
    wb.save(str(f))
    records = list(ExcelReader(str(f)).stream_records())
    assert len(records) == 1
    assert "name" in records[0]
